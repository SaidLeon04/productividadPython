from openpyxl import Workbook
"""
Crea un libro de excel con dos hojas, añade texto a diferentes celdas en ambas hojas

def crear_libro(path):
	workbook = Workbook() # crea un objeto de la clase woorkbook
	sheet = workbook.active # crea la primera hoja
	sheet.title = "productos" " # cambia el nombre de la hoja
	sheet2 = workbook.create_sheet(title =  "tiendas") # crea una segunda hoja y le coloca el nombre en la misma linea

# añade texto a diferentes celdas referenciadas
	sheet["A1"] = "Se viene cositas"
	sheet["A2"] = "Se viene cositas"
	sheet2["A1"] = "Se viene cositas"
	sheet2["A3"] = "Se viene cositas"
	workbook.save(path) 3 guarda los cambios en el libro
	print (workbook.sheetnames) # imprime el nombre de las hojas ya creadas, solo para comprobar


if __name__ == "__main__":
	crear_libro("demo.xlsx") # ejecuta la función, escribe el nombre del archivo/libro
